from openpyxl import load_workbook
from flet import *
import requests
import re
import json
import random, os


class jsonHandler:
    def __init__(self):
        self.submissiosFile = open(
            "src/assets/submissiosFile.json", "r", encoding="utf-8"
        )

    def readFile(self):
        try:
            data = json.load(self.submissiosFile)
            return [True, data]
        except Exception as e:
            return [False, []]


class FileReader:
    def __init__(self, page):
        self.filePath = []
        self.tokin = ""
        self.page = page
        self.rows = []
        self.filteredSubmissionsRow = []
        self.submitionsNamesRow = []
        self.submissionsFromJsonFile = []
        self.submissionFilePath = {}
        self.jsonHandler = jsonHandler()

    def pick_files_result(self, e: FilePickerResultEvent):
        self.filePath = [filePath.path for filePath in e.files] if e.files else []
        if e.files:
            self.disapledButton("start", False)
            self.page.succesTableRef.current.rows = []
            self.page.FiledTableRef.current.rows = []
            for file in self.filePath:
                self.rows.extend(self.readFileRow(file))
                self.submissionsWithCategories = self.readSubmissionsOnly()
                self.submissionsFromJsonFile = self.jsonHandler.readFile()
                self.submitionsNamesRow = self.rows[1]
        else:
            self.page.succesTableRef.current.rows = []
            self.page.FiledTableRef.current.rows = []
            self.disapledButton("start", True)

    def readFileRow(self, filePath):
        if len(self.filePath) and filePath:
            workbook = load_workbook(filePath)
            sheet = workbook.active  # اختيار الورقة النشطة
            rows = []
            for row in sheet.iter_rows(values_only=True):
                row = list(row)
                row.append(filePath)
                rows.append(tuple(row))
            return rows

    def disapledButton(self, buttonName, state):
        if buttonName == "start":
            self.page.startButton.current.disabled = state
            self.page.update()
        elif buttonName == "stop":
            self.page.stopButton.current.disabled = state
            self.page.update()

    def readSubmissionsOnly(self):
        pattern = r"[A-Z]+\d+"
        for row in self.rows:
            if row and isinstance(row[0], str):
                matches = re.findall(pattern, row[0])
                if len(matches):
                    self.filteredSubmissionsRow.append(row)
                    self.submissionFilePath[f"{row[0]}"] = row[-1]
        return self.filteredSubmissionsRow

    def calculatPrice(self, totalOrNot, item, itemPrice):
        if totalOrNot:
            return sum(totalOrNot)
        else:
            return int(item) * int(itemPrice)

    def getCategoryFromRow(self, submission):
        submissionsNameWithValue = {}
        for index, category in enumerate(submission):
            if type(category) == int:
                submissionsNameWithValue[f"{self.submitionsNamesRow[index]}"] = category
        return submissionsNameWithValue

    def genrateHash(self, catrgoryLength=10):
        return random.sample(range(1, 10000), catrgoryLength)

    def getSubmissionsReadyWithCategories(self):
        submissions = {}
        for submission in self.filteredSubmissionsRow:
            submissionData = self.getCategoryFromRow(submission)
            submissions[submission[0]] = {}
            submissions[submission[0]]["additionalSectionSuffixes"] = {
                "67afe8bda9b44d6191471e7e103e8154": []
            }
            hashNumbers = self.genrateHash(len(submission))

            if self.submissionsFromJsonFile[0]:
                for categoryName, itemsNumber in submissionData.items():
                    submissionjson = self.submissionsFromJsonFile[1].get(
                        categoryName, None
                    )
                    if submissionjson:
                        submissionjson["c502ed9800bf4a729eab7572401709ba"] = itemsNumber
                        submissionjson["c63c5bcf0f4d4771ba927c4af3490189"] = str(
                            self.calculatPrice(
                                None,
                                itemsNumber,
                                submissionjson["377e7fcf02894e9aa15c708e5b5180a5"],
                            )
                        )
                        for filedKey, filedValue in submissionjson.items():
                            if filedKey not in submissions[submission[0]].keys():
                                submissions[submission[0]][f"{filedKey}"] = filedValue
                            else:
                                submissions[submission[0]][
                                    f"{filedKey}##{hashNumbers[-1]}"
                                ] = filedValue
                        submissions[submission[0]]["additionalSectionSuffixes"][
                            "67afe8bda9b44d6191471e7e103e8154"
                        ].append(str(hashNumbers[-1]))
                        hashNumbers.pop()
        if submissions.get("Grand Total", None):
            submissions.pop("Grand Total")
        return submissions


class MainCommands:
    def __init__(self, page):
        self.page = page
        self.accesTockin = self.page.client_storage.get("accesss")
        self.accesTockinTextBox = Ref[TextField]()
        self.stopAddSubmissions = False
        self.data = []

    def loaderUi(self):
        self.page.controls.clear()
        self.page.controls.append(
            Column(
                controls=[
                    Container(
                        content=ProgressRing(width=50, height=50, value=None),
                        alignment=alignment.center,
                        expand=True,
                    )
                ],
                expand=True,
                horizontal_alignment=CrossAxisAlignment.CENTER,
                alignment=MainAxisAlignment.CENTER,
            )
        )
        self.page.update()

    def showPersonIcon(self, e):
        self.page.appbar.actions = [
            IconButton(
                icon=Icons.PERSON, on_click=self.page.mainCommandes.addTockinTextBox
            )
        ]
        self.page.update()

    async def sendGetReques(self, subUrl, body={}, params={}):
        headers = {
            "User-Agent": "Mozilla/5.0 (iPhone; CPU iPhone OS 5_1 like Mac OS X) AppleWebKit/534.46 (KHTML, like Gecko) Version/5.1 Mobile/9B176 Safari/7534.48.3",
            "Accept": "application/json, text/plain, */*",
            "Accept-Language": "ar,en-US;q=0.7,en;q=0.3",
            "Accept-Encoding": "gzip, deflate, br",
            "Referer": "https://meter.clappia.com/",
            "Authorization": f"{self.accesTockin}",
            "Origin": "https://meter.clappia.com",
            "Sec-Fetch-Dest": "empty",
            "Sec-Fetch-Mode": "cors",
            "Sec-Fetch-Site": "same-site",
            "Priority": "u=0",
            "Te": "trailers",
        }
        try:
            response = requests.get(
                f"{self.page.baseUrl}/{subUrl}",
                json=body if body else None,
                headers=headers,
                params=params if params else None,
            )
            if response.status_code == 200:
                return [response.status_code, response.json()]
            else:
                return [response.status_code, response.text]

        except Exception as e:
            return [400, e]

    async def sendPostReques(self, subUrl, body={}):
        headers = {
            "Host": "apiv2.clappia.com",
            "User-Agent": "Mozilla/5.0 (iPhone; CPU iPhone OS 5_1 like Mac OS X) AppleWebKit/534.46 (KHTML, like Gecko) Version/5.1 Mobile/9B176 Safari/7534.48.3",
            "Accept": "application/json, text/plain, */*",
            "Accept-Language": "ar,en-US;q=0.7,en;q=0.3",
            "Content-Type": "application/json",
            "Appid": "ALG798812",
            "Workplaceid": "MET860503",
            "Device": "Unknown",
            "Browser": "Firefox",
            "Browserversion": "136.0",
            "Os": "Windows",
            "Platform": "WEB",
            "Osversion": "windows-10",
            "Emailid": "naderalsawahel@gmail.com",
            "Usersignedupat": "1737451321191",
            "Subdomainprefix": "meter",
            "Workplacecreatedat": "1732864785000",
            "Authorization": f"{self.accesTockin}",
        }
        try:
            response =  requests.post(
                f"{self.page.baseUrl}/{subUrl}",
                json=body if body else None,
                headers=headers,
            )
            if response.status_code == 200:
                return [response.status_code, response.json()]
            else:
                return [response.status_code, response.text]

        except Exception as e:
            return [400, e]

    def checkAccesTockin(self, e):
        if self.accesTockinTextBox.current.value:
            params = {
                "appId": "ALG798812",
                "submissionId": "WCK77345648",
                "workplaceId": "MET860503",
                "requestingUserId": "d3eeee5c-7365-4b49-9d1e-7642401d23dd",
            }
            self.accesTockin = self.accesTockinTextBox.current.value
            statusCode, responce = self.page.run_task(
                self.sendGetReques, "appdatav2/getSubmission", {}, params
            ).result()
            if statusCode == 200:
                self.showMessage("تم تخزين المفتاح بنجاح")
                self.showPersonIcon("")
                self.page.client_storage.set("accesss", self.accesTockin)
            else:
                self.showMessage("مفتاح الوصول غير صالح")
        else:
            self.showMessage("الرجاء ادخال مفتاح الوصول اولا")
            self.showMessage("مQفتاح الوصول غير صالح")

    def addTockinTextBox(self, e):
        self.page.appbar.actions = [
            Row(
                controls=[
                    TextField(
                        label="ادخل المفتاح",
                        ref=self.accesTockinTextBox,
                        color="#000000",
                    ),
                    IconButton(icon=Icons.SEND, on_click=self.checkAccesTockin),
                    IconButton(icon=Icons.CLOSE, on_click=self.showPersonIcon),
                ]
            )
        ]
        self.page.update()

    def showMessage(self, text):
        self.page.open(SnackBar(Text(f"{text}"), show_close_icon=True))

    def showResultInPage(self, result , State):
        if State:
            self.page.succesTableRef.current.rows.append(
                DataRow(
                    cells=[
                        DataCell(Text(f"{result['submissionId']}")),
                        DataCell(Text(f"{result['fileName']}")),
                        DataCell(Text(f"{result['path']}")),
                    ],
                ),
            )
            self.page.update()
        else :
            self.page.FiledTableRef.current.rows.append(
                DataRow(
                    cells=[
                        DataCell(Text(f"{result['submissionId']}")),
                        DataCell(Text(f"{result['fileName']}")),
                        DataCell(Text(f"{result['path']}")),
                    ],
                ),
            )
            self.page.update()

    async def getSubmission(self):
        self.data = self.page.fileReader.getSubmissionsReadyWithCategories()
        for key, value in dict(self.data).items():
            print("start")
            params = {
                "appId": "ALG798812",
                "submissionId": key,
                "workplaceId": "MET860503",
                "requestingUserId": "d3eeee5c-7365-4b49-9d1e-7642401d23dd",
            }
            statusCode, responce = await self.sendGetReques("appdatav2/getSubmission", {}, params)
            if statusCode == 200:
                responce = {
                    "submissionFieldValues": responce["submissionFieldValues"],
                    "version": responce["version"],
                    "additionalSectionSuffixes": responce["additionalSectionSuffixes"],
                    "versions": responce["versions"],
                }
                if responce["additionalSectionSuffixes"].get(
                    "67afe8bda9b44d6191471e7e103e8154", None
                ):
                    responce["additionalSectionSuffixes"][
                        "67afe8bda9b44d6191471e7e103e8154"
                    ].extend(
                        value["additionalSectionSuffixes"][
                            "67afe8bda9b44d6191471e7e103e8154"
                        ]
                    )
                else:
                    responce["additionalSectionSuffixes"][
                        "67afe8bda9b44d6191471e7e103e8154"
                    ] = value["additionalSectionSuffixes"][
                        "67afe8bda9b44d6191471e7e103e8154"
                    ]
                value.pop("additionalSectionSuffixes")
                responce["submissionFieldValues"].update(value)

                responce["submissionId"] = key
                responce["workplaceId"] = "MET860503"
                responce["appId"] = "ALG798812"

                responce["isTrialSubmission"] = False

                statusCode, responce_tow = await self.sendPostReques("/appdatav2/editSubmission", responce)
                if statusCode == 200:
                    result = {
                        "submissionId": responce["submissionId"],
                        "path": self.page.fileReader.submissionFilePath.get(
                            responce["submissionId"], None
                        ),
                        "fileName": os.path.basename(
                            self.page.fileReader.submissionFilePath.get(
                                responce["submissionId"], None
                            )
                        ),
                    }
                    self.showResultInPage(result , True)
                else:
                    result = {
                        "submissionId": responce["submissionId"],
                        "path": self.page.fileReader.submissionFilePath.get(
                            responce["submissionId"], None
                        ),
                        "fileName": os.path.basename(
                            self.page.fileReader.submissionFilePath.get(
                                responce["submissionId"], None
                            )
                        ),
                    }
                    self.showResultInPage(result , False)
                self.data.pop(key)
                if self.stopAddSubmissions:
                    break

    def disapledButton(self, buttonName, state):
        if buttonName == "start":
            self.page.startButton.current.disabled = state
            self.page.update()
        elif buttonName == "stop":
            self.page.stopButton.current.disabled = state
            self.page.update()

    def stopAdd(self , e):
        self.disapledButton("stop" , True)
        self.disapledButton("start" , False)
        self.stopAddSubmissions = True
    
    def startAdd(self ,e):
        self.disapledButton("stop" , False)
        self.disapledButton("start" , True)
        self.page.run_task(self.getSubmission)